from __future__ import annotations

import io

from openpyxl import load_workbook

from generate_excel_report import generate_report
from tests.conftest import make_orders_df


def test_generate_report_builds_20_sheet_workbook_from_fixture() -> None:
    report_bytes = generate_report(make_orders_df())
    workbook = load_workbook(io.BytesIO(report_bytes), read_only=True, data_only=True)

    expected_sheets = [
        "1. Summary KPIs",
        "2. Monthly Trend",
        "3. Quarterly Performance",
        "4. Top States",
        "5. Top Cities",
        "6. Top Dishes",
        "7. Category Mix",
        "8. Customer Segments",
        "9. Pareto Analysis",
        "10. Day of Week",
        "11. Time of Day",
        "12. Price-Rating",
        "13. Restaurant Frequency",
        "14. RFM Summary",
        "15. RFM Detail",
        "16. Cohort Retention",
        "17. Statistical Tests",
        "18. Forecast Validation",
        "19. Forecast Metrics",
        "20. Revenue Forecast",
    ]

    assert isinstance(report_bytes, bytes)
    assert workbook.sheetnames == expected_sheets
    assert workbook["1. Summary KPIs"]["A1"].value == "Swiggy Sales Analytics — Executive KPI Summary"
    assert workbook["1. Summary KPIs"]["B5"].value == "6"
    assert workbook["14. RFM Summary"]["A1"].value == "Restaurant-Partner RFM Segment Summary"
    assert workbook["16. Cohort Retention"]["A1"].value == "Monthly Restaurant-Partner Cohort Retention (%)"
    assert workbook["20. Revenue Forecast"]["A1"].value == "Next 3-Month Revenue Forecast"
