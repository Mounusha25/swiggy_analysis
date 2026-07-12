"""
generate_excel_report.py
------------------------
Generates a formatted 20-sheet Excel KPI report from Swiggy order data
using openpyxl.

Standalone usage:
    python3 generate_excel_report.py
    # → writes swiggy_kpi_report.xlsx in the current directory

Imported by app.py to provide a Streamlit download button.
"""

import io
import os
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from analytics_models import (
    calculate_cohort_retention,
    calculate_rfm_segments,
    prepare_order_data,
    run_statistical_tests,
    validate_revenue_forecast,
)

# ---------------------------------------------------------------------------
# Brand palette
# ---------------------------------------------------------------------------
ORANGE = "FF6B35"
DARK = "2D3142"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _sheet_title(ws, row: int, n_cols: int, text: str, bg: str = ORANGE) -> None:
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=13)
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30


def _col_headers(ws, row: int, headers: list) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill(fill_type="solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center")


def _write_df(ws, df: pd.DataFrame, start_row: int) -> None:
    for r_idx, row in enumerate(df.itertuples(index=False), start_row):
        for c_idx, val in enumerate(row, 1):
            # Convert numpy types to native Python for openpyxl compatibility
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = Alignment(horizontal="center")
            if r_idx % 2 == 0:
                c.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GRAY)


def _autofit(ws, df: pd.DataFrame, pad: int = 4) -> None:
    for i, col in enumerate(df.columns, 1):
        header_len = len(str(col))
        data_len = df[col].map(lambda val: len(str(val))).max() if len(df) else 0
        ws.column_dimensions[get_column_letter(i)].width = min(
            max(header_len, data_len) + pad, 42
        )


def _build_sheet(wb: Workbook, sheet_name: str, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(sheet_name)
    _sheet_title(ws, 1, len(df.columns), title)
    _col_headers(ws, 2, df.columns.tolist())
    _write_df(ws, df, 3)
    _autofit(ws, df)


# ---------------------------------------------------------------------------
# Report generator
# ---------------------------------------------------------------------------

def generate_report(df: pd.DataFrame, output_path: str = None):
    """
    Build a 20-sheet Excel KPI workbook from the Swiggy orders DataFrame.

    Parameters
    ----------
    df : pd.DataFrame
        Raw or preprocessed Swiggy order data.
    output_path : str, optional
        If provided, the workbook is saved to this path and the path is returned.
        If None (default), the workbook bytes are returned for Streamlit download.

    Returns
    -------
    str | bytes
        File path (if output_path given) or raw bytes.
    """
    df = prepare_order_data(df, include_synthetic_hour=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ── Sheet 1 : Summary KPIs ─────────────────────────────────────────────
    ws = wb.create_sheet("1. Summary KPIs")
    monthly_rev = df.groupby("Year-Month")["Price (INR)"].sum().sort_index()
    mom = (
        (monthly_rev.iloc[-1] - monthly_rev.iloc[-2]) / monthly_rev.iloc[-2] * 100
        if len(monthly_rev) > 1 else 0.0
    )
    avg_growth = monthly_rev.pct_change().mean() * 100

    _sheet_title(ws, 1, 3, "Swiggy Sales Analytics — Executive KPI Summary")
    kpis = [
        ("KPI", "Value", "Notes"),
        ("Total Revenue (INR)", f"{df['Price (INR)'].sum():,.0f}", "Sum of all order values"),
        ("Total Orders", f"{len(df):,}", "Count of all order records"),
        ("Avg Order Value (INR)", f"{df['Price (INR)'].mean():,.2f}", "Mean basket size"),
        ("Avg Rating", f"{df['Rating'].mean():.2f} / 5.0", "Mean customer satisfaction"),
        ("MoM Growth — Last Month", f"{mom:.2f}%", "Month-over-month revenue change"),
        ("Avg Monthly Growth Rate", f"{avg_growth:.2f}%", "Average MoM growth"),
        ("Unique States", str(df["State"].nunique()), "Geographic coverage"),
        ("Unique Cities", str(df["City"].nunique()), "City-level reach"),
        ("Unique Restaurants", str(df["Restaurant Name"].nunique()), "Partner count"),
        ("Unique Dishes", str(df["Dish Name"].nunique()), "Menu diversity"),
        ("Unique Categories", str(df["Category"].nunique()), "Cuisine diversity"),
        ("Total Rating Count", f"{df['Rating Count'].sum():,}", "Total customer reviews"),
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M"), ""),
    ]
    for r_i, row_data in enumerate(kpis, 3):
        for c_i, val in enumerate(row_data, 1):
            c = ws.cell(row=r_i, column=c_i, value=val)
            if r_i == 3:
                c.font = Font(bold=True, color=WHITE, size=10)
                c.fill = PatternFill(fill_type="solid", fgColor=DARK)
                c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(horizontal="left")
                if r_i % 2 == 0:
                    c.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GRAY)
    for col_letter, width in zip(["A", "B", "C"], [35, 22, 38]):
        ws.column_dimensions[col_letter].width = width

    # ── Sheet 2 : Monthly Trend ────────────────────────────────────────────
    monthly_df = (
        df.groupby("Year-Month")
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Order_Value=("Price (INR)", "mean"),
             Avg_Rating=("Rating", "mean"))
        .round(2)
        .reset_index()
    )
    monthly_df.columns = ["Month", "Orders", "Revenue (INR)", "Avg Order Value (INR)", "Avg Rating"]
    monthly_df["MoM Growth (%)"] = monthly_df["Revenue (INR)"].pct_change().mul(100).round(2)
    _build_sheet(wb, "2. Monthly Trend", "Monthly Revenue Trend & Seasonality", monthly_df)

    # ── Sheet 3 : Quarterly Performance ───────────────────────────────────
    qdf = (
        df.groupby("Quarter")
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Order_Value=("Price (INR)", "mean"),
             Avg_Rating=("Rating", "mean"))
        .round(2)
        .reset_index()
    )
    qdf.columns = ["Quarter", "Orders", "Revenue (INR)", "Avg Order Value (INR)", "Avg Rating"]
    _build_sheet(wb, "3. Quarterly Performance", "Quarterly Sales Performance", qdf)

    # ── Sheet 4 : Top States ──────────────────────────────────────────────
    state_df = (
        df.groupby("State")
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Order_Value=("Price (INR)", "mean"),
             Avg_Rating=("Rating", "mean"),
             Total_Rating_Count=("Rating Count", "sum"))
        .round(2)
        .sort_values("Revenue", ascending=False)
        .reset_index()
    )
    state_df.columns = [
        "State", "Orders", "Revenue (INR)", "Avg Order Value (INR)",
        "Avg Rating", "Total Rating Count"
    ]
    _build_sheet(wb, "4. Top States", "State-wise Revenue Performance", state_df)

    # ── Sheet 5 : Top Cities ──────────────────────────────────────────────
    city_df = (
        df.groupby(["City", "State"])
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Order_Value=("Price (INR)", "mean"),
             Avg_Rating=("Rating", "mean"))
        .round(2)
        .sort_values("Revenue", ascending=False)
        .head(30)
        .reset_index()
    )
    city_df.columns = ["City", "State", "Orders", "Revenue (INR)", "Avg Order Value (INR)", "Avg Rating"]
    _build_sheet(wb, "5. Top Cities", "Top 30 Cities by Revenue", city_df)

    # ── Sheet 6 : Top Dishes ──────────────────────────────────────────────
    dish_df = (
        df.groupby("Dish Name")
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Price=("Price (INR)", "mean"),
             Avg_Rating=("Rating", "mean"))
        .round(2)
        .sort_values("Revenue", ascending=False)
        .head(30)
        .reset_index()
    )
    dish_df.columns = ["Dish Name", "Orders", "Revenue (INR)", "Avg Price (INR)", "Avg Rating"]
    _build_sheet(wb, "6. Top Dishes", "Top 30 Dishes by Revenue", dish_df)

    # ── Sheet 7 : Category Mix ────────────────────────────────────────────
    cat_df = (
        df.groupby("Category")
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Price=("Price (INR)", "mean"),
             Avg_Rating=("Rating", "mean"))
        .round(2)
        .sort_values("Revenue", ascending=False)
        .reset_index()
    )
    cat_df.columns = ["Category", "Orders", "Revenue (INR)", "Avg Price (INR)", "Avg Rating"]
    cat_df["Revenue Share (%)"] = (
        cat_df["Revenue (INR)"] / cat_df["Revenue (INR)"].sum() * 100
    ).round(2)
    _build_sheet(wb, "7. Category Mix", "Food Category Revenue Mix", cat_df)

    # ── Sheet 8 : Customer Segments ───────────────────────────────────────
    seg_df = (
        df.groupby("Value_Segment", observed=True)
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Price=("Price (INR)", "mean"),
             Avg_Rating=("Rating", "mean"))
        .round(2)
        .reset_index()
    )
    seg_df.columns = ["Segment", "Orders", "Revenue (INR)", "Avg Order Value (INR)", "Avg Rating"]
    seg_df["Revenue Share (%)"] = (seg_df["Revenue (INR)"] / seg_df["Revenue (INR)"].sum() * 100).round(2)
    seg_df["Order Share (%)"] = (seg_df["Orders"] / seg_df["Orders"].sum() * 100).round(2)
    _build_sheet(wb, "8. Customer Segments", "Customer Segmentation by Basket Value", seg_df)

    # ── Sheet 9 : Pareto Analysis ─────────────────────────────────────────
    pareto_df = (
        df.groupby("City")["Price (INR)"].sum()
        .sort_values(ascending=False)
        .reset_index()
    )
    pareto_df.columns = ["City", "Revenue (INR)"]
    pareto_df["Cumulative Revenue (INR)"] = pareto_df["Revenue (INR)"].cumsum()
    pareto_df["Cumulative % of Revenue"] = (
        pareto_df["Cumulative Revenue (INR)"] / pareto_df["Revenue (INR)"].sum() * 100
    ).round(2)
    pareto_df["City Rank"] = range(1, len(pareto_df) + 1)
    n80 = (pareto_df["Cumulative % of Revenue"] <= 80).sum()
    _build_sheet(
        wb, "9. Pareto Analysis",
        f"Pareto Analysis — {n80} cities drive 80% of total revenue",
        pareto_df,
    )

    # ── Sheet 10 : Day of Week ────────────────────────────────────────────
    day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    dow_df = (
        df.groupby("DayName")
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Order_Value=("Price (INR)", "mean"),
             Avg_Rating=("Rating", "mean"))
        .round(2)
        .reindex(day_order)
        .reset_index()
    )
    dow_df.columns = ["Day of Week", "Orders", "Revenue (INR)", "Avg Order Value (INR)", "Avg Rating"]
    _build_sheet(wb, "10. Day of Week", "Sales Patterns by Day of Week", dow_df)

    # ── Sheet 11 : Time of Day ────────────────────────────────────────────
    slot_order = ["Morning", "Lunch", "Afternoon", "Dinner", "Night", "Late Night"]
    tod_df = (
        df.groupby("Time Slot", observed=True)
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Order_Value=("Price (INR)", "mean"),
             Avg_Rating=("Rating", "mean"))
        .round(2)
        .reindex(slot_order)
        .reset_index()
    )
    tod_df.columns = ["Time Slot", "Orders", "Revenue (INR)", "Avg Order Value (INR)", "Avg Rating"]
    tod_df["Order Share (%)"] = (tod_df["Orders"] / tod_df["Orders"].sum() * 100).round(2)
    _build_sheet(
        wb, "11. Time of Day",
        "Order Patterns by Time of Day (modelled distribution)",
        tod_df,
    )

    # ── Sheet 12 : Price-Rating Correlation ──────────────────────────────
    corr = df["Price (INR)"].corr(df["Rating"])
    price_bins = pd.cut(df["Price (INR)"], bins=10)
    corr_df = (
        df.groupby(price_bins, observed=True)
        .agg(Orders=("Price (INR)", "count"),
             Avg_Rating=("Rating", "mean"),
             Avg_Price=("Price (INR)", "mean"))
        .round(3)
        .reset_index()
    )
    corr_df.columns = ["Price Range", "Orders", "Avg Rating", "Avg Price (INR)"]
    corr_df["Price Range"] = corr_df["Price Range"].astype(str)
    _build_sheet(
        wb, "12. Price-Rating",
        f"Price vs Rating Correlation  (Pearson r = {corr:.3f})",
        corr_df,
    )

    # ── Sheet 13 : Restaurant Frequency Tiers ────────────────────────────
    rest_df = (
        df.groupby("Restaurant Name")
        .agg(Orders=("Price (INR)", "count"),
             Revenue=("Price (INR)", "sum"),
             Avg_Rating=("Rating", "mean"),
             City=("City", "first"),
             State=("State", "first"))
        .round(2)
        .sort_values("Orders", ascending=False)
        .reset_index()
    )
    rest_df["Frequency Tier"] = pd.cut(
        rest_df["Orders"].rank(pct=True),
        bins=[0, 0.40, 0.80, 1.0],
        labels=["Low Volume (Bottom 40%)", "Medium Volume (40-80%)", "High Volume (Top 20%)"],
    )
    rest_df = rest_df[
        ["Restaurant Name", "City", "State", "Frequency Tier", "Orders", "Revenue", "Avg_Rating"]
    ]
    rest_df.columns = [
        "Restaurant", "City", "State", "Frequency Tier",
        "Orders", "Revenue (INR)", "Avg Rating"
    ]
    _build_sheet(
        wb, "13. Restaurant Frequency",
        "Restaurant Order Frequency Segmentation (Top 100)",
        rest_df.head(100),
    )

    # ── Sheet 14 : RFM Summary ────────────────────────────────────────────
    rfm_df, rfm_summary = calculate_rfm_segments(df)
    rfm_summary = rfm_summary.rename(
        columns={
            "Entities": "Restaurants",
            "Avg_Recency_Days": "Avg Recency Days",
            "Avg_Frequency": "Avg Frequency",
            "Total_Revenue": "Total Revenue (INR)",
            "Avg_RFM_Score": "Avg RFM Score",
        }
    )
    _build_sheet(
        wb,
        "14. RFM Summary",
        "Restaurant-Partner RFM Segment Summary",
        rfm_summary,
    )

    # ── Sheet 15 : RFM Detail ─────────────────────────────────────────────
    rfm_detail = rfm_df[
        [
            "Restaurant Name",
            "RFM_Segment",
            "RFM_Code",
            "RFM_Score",
            "Recency_Days",
            "Frequency",
            "Monetary",
            "Avg_Order_Value",
            "Avg_Rating",
        ]
    ].head(250)
    rfm_detail.columns = [
        "Restaurant",
        "RFM Segment",
        "RFM Code",
        "RFM Score",
        "Recency Days",
        "Frequency",
        "Revenue (INR)",
        "Avg Order Value (INR)",
        "Avg Rating",
    ]
    _build_sheet(
        wb,
        "15. RFM Detail",
        "Top 250 Restaurants by RFM Score",
        rfm_detail,
    )

    # ── Sheet 16 : Cohort Retention ───────────────────────────────────────
    cohort_df = calculate_cohort_retention(df)
    _build_sheet(
        wb,
        "16. Cohort Retention",
        "Monthly Restaurant-Partner Cohort Retention (%)",
        cohort_df,
    )

    # ── Sheet 17 : Statistical Tests ──────────────────────────────────────
    tests_df = run_statistical_tests(df)
    _build_sheet(
        wb,
        "17. Statistical Tests",
        "Mann-Whitney U and ANOVA Test Results",
        tests_df,
    )

    # ── Sheet 18-20 : Forecast Validation ─────────────────────────────────
    validation_df, forecast_metrics, forecast_df = validate_revenue_forecast(df)
    if validation_df.empty:
        validation_df = pd.DataFrame(
            [{"Status": "Not enough monthly history to validate forecast"}]
        )
        forecast_metrics = pd.DataFrame(
            [{"Status": "Not enough monthly history to calculate MAPE/RMSE"}]
        )
        forecast_df = pd.DataFrame(
            [{"Status": "Not enough monthly history to forecast"}]
        )

    _build_sheet(
        wb,
        "18. Forecast Validation",
        "Holdout Forecast Validation by Model",
        validation_df,
    )
    _build_sheet(
        wb,
        "19. Forecast Metrics",
        "Forecast Accuracy: MAPE and RMSE",
        forecast_metrics,
    )
    _build_sheet(
        wb,
        "20. Revenue Forecast",
        "Next 3-Month Revenue Forecast",
        forecast_df,
    )

    # ── Output ────────────────────────────────────────────────────────────
    if output_path:
        wb.save(output_path)
        return output_path
    else:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()


# ---------------------------------------------------------------------------
# Standalone execution
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("Reading swiggy_data.xlsx …")
    raw_df = pd.read_excel("swiggy_data.xlsx")
    out_path = "swiggy_kpi_report.xlsx"
    generate_report(raw_df, output_path=out_path)
    size_kb = os.path.getsize(out_path) // 1024
    print(f"Excel report saved: {out_path}  ({size_kb} KB, 20 sheets)")
